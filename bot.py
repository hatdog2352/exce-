import os
import re
import shutil
import copy
import unicodedata
from datetime import datetime
from pathlib import Path

import discord
from discord.ext import commands
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Alignment, Protection
from dotenv import load_dotenv

load_dotenv()

TOKEN = os.getenv("DISCORD_TOKEN", "").strip()
HOUSE_CHANNEL_ID = int(os.getenv("HOUSE_CHANNEL_ID", "0") or 0)

BASE_DIR = Path(__file__).resolve().parent
TEMPLATE_DIR = BASE_DIR / "templates"
TEMPLATE_PATH = TEMPLATE_DIR / "template.xlsx"
OUTPUT_DIR = BASE_DIR / "output"
OUTPUT_DIR.mkdir(exist_ok=True)
TEMPLATE_DIR.mkdir(exist_ok=True)

if not TOKEN:
    raise RuntimeError("DISCORD_TOKEN is missing from .env")

intents = discord.Intents.default()
intents.message_content = True
intents.messages = True
intents.guilds = True

bot = commands.Bot(command_prefix="!", intents=intents)

PRICE_RE = re.compile(r"(?:₱\s*|peso\s*)?(\d+)\s*(?:pesos?|peso)?", re.I)
NUMBERED_RE = re.compile(
    r"^\s*(\d+)\s*[\.\)]?\s*(.*?)\s*(?:—|–|-)\s*(?:₱\s*)?(\d+)\s*$"
)

URL_RE = re.compile(r"https?://\S+", re.I)

# Ordered longest-first so "queenslander house" wins over "house".
TYPE_PATTERNS = [
    ("Queenslander House", [r"\bqueenslander\s+house\b", r"\bqueenslander\b"]),
    ("Gingerbread House", [r"\bgingerbread\s+house\b", r"\bgingerbread\b"]),
    ("Tiny Home", [r"\btiny\s+home\b", r"\btiny\s+homes\b"]),
    ("Tiny House", [r"\btiny\s+house\b", r"\btiny\s+houses\b"]),
    ("Bunker House", [r"\bbunker\s+house\b", r"\bbunker\b"]),
    ("Family Home", [r"\bfamily\s+home\b", r"\bfamily\s+house\b"]),
    ("Family House", [r"\bfamily\s+house\b"]),
    ("Treehouse", [r"\btreehouse\b", r"\btree\s*house\b"]),
    ("Estate House", [r"\bestate\s+house\b"]),
    ("Mountain House", [r"\bmountain\s+house\b"]),
    ("Racetrack House", [r"\bracetrack\s+house\b", r"\bracetrack\b"]),
    ("Sandbox Island", [r"\bsandbox\s+island\b"]),
    ("Hollywood House", [r"\bhollywood\s+house\b"]),
    ("Pizzaplace House", [r"\bpizzaplace\s+house\b", r"\bpizzaplace\b"]),
    ("Shop House", [r"\bshop\s+house\b"]),
    ("Safari House", [r"\bsafari\s+house\b"]),
    ("Friendly House", [r"\bfriendly\s+house\b"]),
]

HEADER_RE = re.compile(r"(\d+)\s*(?:peso|pesos)\s*houses?", re.I)


def normalize_digits(text: str) -> str:
    out = []
    for ch in text:
        try:
            out.append(str(unicodedata.digit(ch)))
        except (TypeError, ValueError):
            out.append(ch)
    return "".join(out)


def clean_line(line: str) -> str:
    line = normalize_digits(line)
    line = re.sub(r"[*_`~]", "", line)
    line = line.replace("\u200b", "").replace("\xa0", " ")
    return line.strip()


def extract_price(text: str):
    text = normalize_digits(text)
    m = re.search(r"₱\s*(\d+)", text)
    if m:
        return int(m.group(1))
    m = HEADER_RE.search(text)
    if m:
        return int(m.group(1))
    return None


def extract_type(lines):
    # Explicit "House Type / Cost: queenslander house / 53k"
    for line in lines:
        m = re.search(
            r"house\s*type\s*(?:/\s*cost)?\s*:\s*(.+?)(?:\s*/\s*[\d,.]+k?)?\s*$",
            line,
            re.I,
        )
        if m:
            value = m.group(1).strip()
            # Remove trailing cost if regex did not consume it.
            value = re.sub(r"\s*/\s*[\d,.]+k?\s*$", "", value, flags=re.I)
            t = canonical_type(value)
            if t:
                return t

    for line in lines:
        t = canonical_type(line)
        if t:
            return t

    return ""


def canonical_type(text: str):
    text = text.strip()
    for canonical, patterns in TYPE_PATTERNS:
        for pattern in patterns:
            if re.search(pattern, text, re.I):
                return canonical
    return ""


def extract_file(lines):
    # Explicit "File name: something"
    for line in lines:
        m = re.match(r"^\s*file\s*name\s*:\s*(.+?)\s*$", line, re.I)
        if m:
            return m.group(1).strip()

    # "cutecoress file name"
    for line in lines:
        m = re.match(r"^\s*(.+?)\s+file\s*name\s*$", line, re.I)
        if m:
            return m.group(1).strip()

    # URL / Pastebin
    for line in lines:
        m = URL_RE.search(line)
        if m:
            return m.group(0).rstrip(").,]")

    # Last useful line after excluding obvious house-type/cost lines.
    candidates = []
    for line in lines:
        low = line.lower().strip()
        if not low:
            continue
        if re.match(r"^(house\s*type|file\s*name)\s*:", low):
            continue
        if canonical_type(line):
            # "52k queenslander" is type/cost, not the file.
            if re.search(r"\b\d+(?:\.\d+)?k\b", line, re.I):
                continue
            continue
        if re.fullmatch(r"\d+(?:\.\d+)?k(?:\+)?", low, re.I):
            continue
        candidates.append(line)

    return candidates[-1].strip() if candidates else ""


def parse_house_message(content: str):
    raw_lines = [clean_line(x) for x in content.splitlines()]
    lines = [x for x in raw_lines if x]

    entries = []
    current_price = None
    current = None

    def finish():
        nonlocal current
        if not current:
            return

        details = current["details"]
        price = current["price"] or current_price
        house_type = extract_type(details)
        file_name = extract_file(details)

        # If no explicit type was present, use a conservative default.
        # This is preferable to putting cost text into the House Types column.
        if not house_type:
            house_type = "Tiny Home"

        entries.append({
            "number": current["number"],
            "name": current["name"].strip(),
            "price": price,
            "file": file_name,
            "type": house_type,
        })
        current = None

    for line in lines:
        # Price section heading, e.g. "✨🌷 𝟱𝟱 PESO HOUSES 🌷✨"
        header_price = extract_price(line) if HEADER_RE.search(normalize_digits(line)) else None
        if header_price is not None and "house" in line.lower():
            finish()
            current_price = header_price
            continue

        m = NUMBERED_RE.match(line)
        if m:
            finish()
            number = int(m.group(1))
            name = m.group(2).strip()
            price = int(m.group(3))
            current = {
                "number": number,
                "name": name,
                "price": price,
                "details": [],
            }
            current_price = price
            continue

        if current is not None:
            # Ignore separator/header lines inside a block.
            if HEADER_RE.search(normalize_digits(line)):
                continue
            current["details"].append(line)

    finish()

    # Remove malformed entries and group by price.
    entries = [e for e in entries if e["name"] and e["price"] is not None]
    grouped = {}
    for e in entries:
        grouped.setdefault(e["price"], []).append(e)

    return grouped


def copy_row_style(ws, source_row: int, target_row: int, start_col: int, end_col: int):
    if source_row < 1 or source_row > ws.max_row:
        return
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height

    for col in range(start_col, end_col + 1):
        src = ws.cell(source_row, col)
        dst = ws.cell(target_row, col)

        if src.has_style:
            dst._style = copy.copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        dst.alignment = copy.copy(src.alignment)
        dst.protection = copy.copy(src.protection)


def is_price_header(value):
    if not isinstance(value, str):
        return None
    normalized = normalize_digits(value)
    m = HEADER_RE.search(normalized)
    return int(m.group(1)) if m else None


def find_price_sections(ws):
    sections = []
    for start_col in (1, 5):
        for row in range(1, ws.max_row + 1):
            price = is_price_header(ws.cell(row, start_col).value)
            if price is not None:
                sections.append({
                    "price": price,
                    "start_col": start_col,
                    "end_col": start_col + 2,
                    "header_row": row,
                })

    # Find the next header on the SAME side.
    for s in sections:
        next_rows = [
            x["header_row"]
            for x in sections
            if x["start_col"] == s["start_col"]
            and x["header_row"] > s["header_row"]
        ]
        s["next_header_row"] = min(next_rows) if next_rows else None

    return sections


def find_section(ws, price):
    sections = [s for s in find_price_sections(ws) if s["price"] == price]
    if not sections:
        return None

    # Normally there is exactly one. If duplicated, prefer the one with
    # more existing data rows.
    def capacity(s):
        end = s["next_header_row"] or (ws.max_row + 1)
        return max(0, end - (s["header_row"] + 2))
    return max(sections, key=capacity)


def clear_side(ws, start_row, end_row, start_col, end_col):
    if end_row < start_row:
        return
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            ws.cell(row, col).value = None


def write_section(ws, section, entries):
    start_col = section["start_col"]
    end_col = section["end_col"]
    header_row = section["header_row"]
    next_header = section["next_header_row"]

    first_data_row = header_row + 2

    if next_header is not None:
        available = max(0, next_header - first_data_row)
    else:
        # Do not use the other side's content as a reason to clear rows.
        # We only need enough rows for the new entries.
        available = max(0, ws.max_row - first_data_row + 1)

    needed = len(entries)

    # If the section is too small, insert rows immediately before its next header.
    if needed > available:
        extra = needed - available
        insert_at = next_header if next_header is not None else ws.max_row + 1

        # Copy style from the nearest existing data row before inserting.
        style_source = max(first_data_row, insert_at - 1)
        for i in range(extra):
            ws.insert_rows(insert_at + i, 1)
            copy_row_style(ws, style_source, insert_at + i, start_col, end_col)

        # Recalculate because row insertion changes following headers.
        section = find_section(ws, section["price"])
        start_col = section["start_col"]
        end_col = section["end_col"]
        header_row = section["header_row"]
        next_header = section["next_header_row"]
        first_data_row = header_row + 2

    # Clear the whole available region on this side only.
    if next_header is not None:
        clear_end = next_header - 1
    else:
        clear_end = max(ws.max_row, first_data_row + needed - 1)

    clear_side(ws, first_data_row, clear_end, start_col, end_col)

    # Choose a style source from an existing data row if possible.
    style_source = first_data_row
    if style_source > ws.max_row:
        style_source = header_row + 1
    if ws.cell(style_source, start_col).has_style is False and style_source > 1:
        style_source -= 1

    for i, entry in enumerate(entries):
        row = first_data_row + i
        if row > ws.max_row:
            ws.insert_rows(row, 1)
        copy_row_style(ws, style_source, row, start_col, end_col)

        ws.cell(row, start_col).value = f"{entry['number']}. {entry['name']}"
        ws.cell(row, start_col + 1).value = entry["file"]
        ws.cell(row, start_col + 2).value = entry["type"]


def create_missing_section(ws, price, entries):
    # Put a new section on the side that currently ends higher/has less content.
    left_last = max(
        (r for r in range(1, ws.max_row + 1)
         if any(ws.cell(r, c).value is not None for c in range(1, 4))),
        default=1
    )
    right_last = max(
        (r for r in range(1, ws.max_row + 1)
         if any(ws.cell(r, c).value is not None for c in range(5, 8))),
        default=1
    )

    start_col = 1 if left_last <= right_last else 5
    header_row = max(left_last if start_col == 1 else right_last, 1) + 2

    # Copy a known price-header style and data-row style from the same side.
    source_header = None
    source_data = None
    for s in find_price_sections(ws):
        if s["start_col"] == start_col:
            source_header = s["header_row"]
            source_data = s["header_row"] + 2
            break

    if source_header is None:
        source_header = 1
        source_data = 3

    for col in range(start_col, start_col + 3):
        copy_row_style(ws, source_header, header_row, col, col)
    ws.cell(header_row, start_col).value = f"🌷✨ {price} peso houses 🌷✨"

    data_start = header_row + 2
    for i in range(len(entries)):
        copy_row_style(ws, source_data, data_start + i, start_col, start_col + 2)

    new_section = {
        "price": price,
        "start_col": start_col,
        "end_col": start_col + 2,
        "header_row": header_row,
        "next_header_row": None,
    }
    write_section(ws, new_section, entries)


def update_workbook(grouped):
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(
            f"Template not found at {TEMPLATE_PATH}. Use !settemplate first."
        )

    wb = load_workbook(TEMPLATE_PATH)
    ws = wb["Sheet1"]

    # Process prices from high to low. This makes row insertion less surprising.
    for price in sorted(grouped.keys(), reverse=True):
        section = find_section(ws, price)
        if section:
            write_section(ws, section, grouped[price])
        else:
            create_missing_section(ws, price, grouped[price])

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output = OUTPUT_DIR / f"houses_{timestamp}.xlsx"
    wb.save(output)
    return output


def format_summary(grouped):
    total = sum(len(v) for v in grouped.values())
    lines = [f"✅ Parsed **{total} house(s)**."]
    for price in sorted(grouped):
        lines.append(f"• ₱{price}: {len(grouped[price])} house(s)")
    return "\n".join(lines)


@bot.event
async def on_ready():
    print(f"Logged in as {bot.user} (ID: {bot.user.id})")
    try:
        synced = await bot.tree.sync()
        print(f"Synced {len(synced)} slash command(s).")
    except Exception as e:
        print(f"Slash command sync failed: {e}")


@bot.tree.command(name="settemplate", description="Upload and save the Excel template.")
async def settemplate(interaction: discord.Interaction, file: discord.Attachment):
    if not file.filename.lower().endswith((".xlsx", ".xlsm")):
        await interaction.response.send_message(
            "❌ Please upload an `.xlsx` or `.xlsm` Excel file.",
            ephemeral=True,
        )
        return

    await interaction.response.defer(ephemeral=True)

    try:
        data = await file.read()
        TEMPLATE_PATH.write_bytes(data)

        # Validate Sheet1 before accepting it.
        wb = load_workbook(TEMPLATE_PATH, read_only=True)
        if "Sheet1" not in wb.sheetnames:
            TEMPLATE_PATH.unlink(missing_ok=True)
            await interaction.followup.send(
                "❌ The uploaded workbook does not contain a `Sheet1` sheet.",
                ephemeral=True,
            )
            return
        wb.close()

        await interaction.followup.send(
            f"✅ Template saved as `{TEMPLATE_PATH.name}`.",
            ephemeral=True,
        )
    except Exception as e:
        await interaction.followup.send(
            f"❌ Could not save the template: `{type(e).__name__}: {e}`",
            ephemeral=True,
        )


@bot.tree.command(name="template", description="Check whether the Excel template is installed.")
async def template_status(interaction: discord.Interaction):
    if TEMPLATE_PATH.exists():
        await interaction.response.send_message(
            "✅ Excel template is installed and ready."
        )
    else:
        await interaction.response.send_message(
            "❌ No template found. Use `/settemplate` and upload your Excel template."
        )


@bot.tree.command(name="generate", description="Generate Excel from text attached to this command.")
async def generate(interaction: discord.Interaction, text: str):
    if not TEMPLATE_PATH.exists():
        await interaction.response.send_message(
            "❌ No template found. Use `/settemplate` first.",
            ephemeral=True,
        )
        return

    await interaction.response.defer()

    grouped = parse_house_message(text)
    if not grouped:
        await interaction.followup.send(
            "❌ I couldn't find any numbered house listings with prices."
        )
        return

    try:
        output = update_workbook(grouped)
        await interaction.followup.send(
            content=format_summary(grouped),
            file=discord.File(output, filename=output.name),
        )
    except Exception as e:
        await interaction.followup.send(
            f"❌ Excel generation failed: `{type(e).__name__}: {e}`"
        )


@bot.event
async def on_message(message: discord.Message):
    if message.author.bot:
        return

    # Keep normal prefix commands working.
    await bot.process_commands(message)

    if HOUSE_CHANNEL_ID and message.channel.id != HOUSE_CHANNEL_ID:
        return

    content = message.content.strip()
    if not content:
        return

    # Only auto-process messages that look like the house-list format.
    normalized = normalize_digits(content)
    has_price_header = bool(HEADER_RE.search(normalized))
    has_numbered_house = bool(
        re.search(r"^\s*\d+\s*[\.\)]?.+?(?:—|–|-)\s*(?:₱\s*)?\d+", normalized, re.M)
    )

    if not (has_price_header and has_numbered_house):
        return

    grouped = parse_house_message(content)
    if not grouped:
        return

    try:
        async with message.channel.typing():
            output = update_workbook(grouped)

        await message.reply(
            content=format_summary(grouped),
            file=discord.File(output, filename=output.name),
            mention_author=False,
        )
    except Exception as e:
        await message.reply(
            f"❌ Excel generation failed: `{type(e).__name__}: {e}`",
            mention_author=False,
        )


if __name__ == "__main__":
    bot.run(TOKEN)
